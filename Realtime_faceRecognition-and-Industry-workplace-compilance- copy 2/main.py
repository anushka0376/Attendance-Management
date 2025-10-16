
import cv2
import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import sqlite3
from openpyxl import Workbook, load_workbook
import socket
import warnings
import platform

# Try to import optional dependencies
try:
    import face_recognition
    FACE_RECOGNITION_AVAILABLE = True
    print("✅ Face recognition module loaded")
except ImportError:
    FACE_RECOGNITION_AVAILABLE = False
    print("⚠️ Face recognition module not available - install with: pip install face-recognition")

try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False
    print("⚠️ Google Sheets modules not available")

warnings.filterwarnings("ignore")

# =============================
# CONFIG
# =============================
IMAGE_FOLDER = "images"
UNKNOWN_FOLDER = "unknown_faces"
DB_PATH = "attendance.db"
EXCEL_BACKUP_FILE = "attendance_backup.xlsx"  # one file with daily sheets
# Prefer an environment variable (works cross-platform). Fall back to a file in the
# project root. Use expanduser to allow ~/ paths on macOS/Linux.
SERVICE_ACCOUNT_JSON = os.environ.get(
    "SERVICE_ACCOUNT_JSON",
    os.path.expanduser(os.path.join(os.getcwd(), "attendancesystem-456710-b95819cc04c5.json")),
)
GOOGLE_SHEET_KEY = "146GSoCK3fWDpWrfOTq4XWzuf09_wMnKeueoSmoXraE4"
TARGET_GID = 1841396514
COOLDOWN_SECONDS = 30  # avoid repeat marks for same student within this window

# On macOS/Continuity Camera the frame can be sideways; rotate there by default.
IS_MAC = platform.system() == "Darwin"
# Rotate only on macOS Continuity Camera by default. Keep configurable.
ROTATE_ON_CAPTURE = IS_MAC
ROTATE_CODE = cv2.ROTATE_90_CLOCKWISE

# =============================
# INTERNET CHECK
# =============================
def is_connected(host="8.8.8.8", port=53, timeout=3):
    try:
        socket.setdefaulttimeout(timeout)
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((host, port))
        s.close()
        return True
    except Exception:
        return False

# =============================
# GOOGLE SHEETS SETUP
# =============================
internet_available = is_connected()
sheet = None
worksheet = None

if GSHEETS_AVAILABLE:
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive",
    ]

    if internet_available:
        # Ensure the service account json exists before trying to authenticate.
        if SERVICE_ACCOUNT_JSON and os.path.exists(SERVICE_ACCOUNT_JSON):
            try:
                creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_JSON, scope)
                client = gspread.authorize(creds)
                sheet = client.open_by_key(GOOGLE_SHEET_KEY)
                try:
                    worksheet = sheet.get_worksheet_by_id(TARGET_GID)
                except Exception:
                    worksheet = sheet.get_worksheet(0)
                print("✅ Google Sheets connected.")
            except Exception as e:
                print(f"⚠️ Could not set up Google Sheets: {e}")
                internet_available = False
        else:
            print(f"⚠️ Google service JSON not found at {SERVICE_ACCOUNT_JSON}; Google Sheets disabled.")
            internet_available = False
    else:
        print("⚠️ No internet — Google Sheets disabled for this run.")
else:
    print("⚠️ Google Sheets modules not available - Google Sheets disabled.")

# =============================
# DATABASE SETUP
# =============================
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS departments (
        department_id INTEGER PRIMARY KEY AUTOINCREMENT,
        department_name TEXT UNIQUE
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS semesters (
        semester_id INTEGER PRIMARY KEY AUTOINCREMENT,
        semester_name TEXT UNIQUE
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS groups (
        group_id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_name TEXT UNIQUE
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS students (
        student_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        roll_no TEXT UNIQUE,
        class_name TEXT,
        semester_id INTEGER,
        group_id INTEGER,
        department_id INTEGER,
        FOREIGN KEY (semester_id) REFERENCES semesters(semester_id),
        FOREIGN KEY (group_id) REFERENCES groups(group_id),
        FOREIGN KEY (department_id) REFERENCES departments(department_id)
    )
    """
)
cursor.execute(
    """
    CREATE TABLE IF NOT EXISTS attendance (
        attendance_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        date TEXT,
        day TEXT,
        entry_time TEXT,
        exit_time TEXT,
        timestamp TEXT,
        FOREIGN KEY (student_id) REFERENCES students(student_id)
    )
    """
)
conn.commit()

# =============================
# GLOBALS
# =============================
os.makedirs(IMAGE_FOLDER, exist_ok=True)
os.makedirs(UNKNOWN_FOLDER, exist_ok=True)

known_face_encodings = []
known_face_ids = {}  # index -> student_id
_last_seen = {}      # student_id -> datetime

# =============================
# CAMERA HELPERS
# =============================
def open_camera(preferred_index=0, max_tested=5):
    print("🔍 Testing available cameras...")
    found_any = False
    for i in range(preferred_index, max_tested):
        cap = cv2.VideoCapture(i)
        if cap.isOpened():
            print(f"✅ Camera index {i} is available.")
            found_any = True
            cap.release()
        else:
            print(f"❌ Camera index {i} is not available.")
    
    if not found_any:
        print("⚠️ No cameras found! Please check:")
        print("  - Camera is connected and not in use by another app")
        print("  - Camera drivers are installed")
        print("  - Camera permissions are granted")
        return None

    print(f"🎯 Attempting to open camera {preferred_index}...")
    cap = cv2.VideoCapture(preferred_index)
    if cap.isOpened():
        # Test if we can actually read a frame
        ret, test_frame = cap.read()
        if ret and test_frame is not None:
            print(f"✅ Camera {preferred_index} opened successfully!")
            cv2.namedWindow("Camera Test", cv2.WINDOW_NORMAL)
            return cap
        else:
            print(f"❌ Camera {preferred_index} opened but cannot read frames")
            cap.release()
            return None
    else:
        print(f"❌ Failed to open camera {preferred_index}")
        return None

def maybe_rotate(frame):
    if ROTATE_ON_CAPTURE and frame is not None:
        return cv2.rotate(frame, ROTATE_CODE)
    return frame

# =============================
# LOOKUPS / HELPERS
# =============================
def get_or_create_group(group_name: str) -> int:
    cursor.execute("INSERT OR IGNORE INTO groups (group_name) VALUES (?)", (group_name,))
    conn.commit()
    return cursor.execute("SELECT group_id FROM groups WHERE group_name=?", (group_name,)).fetchone()[0]

def get_or_create_semester(sem_name: str) -> int:
    cursor.execute("INSERT OR IGNORE INTO semesters (semester_name) VALUES (?)", (sem_name,))
    conn.commit()
    return cursor.execute("SELECT semester_id FROM semesters WHERE semester_name=?", (sem_name,)).fetchone()[0]

def get_or_create_department(dept_name: str) -> int:
    cursor.execute("INSERT OR IGNORE INTO departments (department_name) VALUES (?)", (dept_name,))
    conn.commit()
    return cursor.execute("SELECT department_id FROM departments WHERE department_name=?", (dept_name,)).fetchone()[0]

# =============================
# EXCEL HELPERS
# =============================
def ensure_backup_book_and_sheet(date_str: str):
    """Open or create the Excel backup workbook and return (wb, worksheet_for_date).

    This is used to keep a local Excel copy of daily attendance and works on
    Windows and macOS because it uses openpyxl only (no shell calls).
    """
    if os.path.exists(EXCEL_BACKUP_FILE):
        wb = load_workbook(EXCEL_BACKUP_FILE)
    else:
        wb = Workbook()
    # Remove default empty sheet if still present
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        wb.remove(wb["Sheet"])
    if date_str in wb.sheetnames:
        ws = wb[date_str]
    else:
        ws = wb.create_sheet(title=date_str)
        ws.append([
            "Name", "Roll No", "Subject", "Date", "Day",
            "Entry Time", "Exit Time", "Timestamp",
            "Department", "Semester", "Group"
        ])
    return wb, ws

from datetime import datetime

def log_to_google_sheet(sheet, class_tab_base, row_data):
    # Current date string (e.g., "2025-09-28")
    date_str = datetime.now().strftime("%Y-%m-%d")

    # 1. Always append to Backup_Log
    try:
        backup_ws = sheet.worksheet("Backup_Log")
    except gspread.exceptions.WorksheetNotFound:
        backup_ws = sheet.add_worksheet(title="Backup_Log", rows="1000", cols="20")
        backup_ws.append_row([
            "name", "roll_no", "subject_name", "date", "day",
            "entry_time", "exit_time", "timestamp",
            "department_name", "semester_name", "group_name"
        ])
    backup_ws.append_row(row_data)

    # 2. Create/find today's class tab
    today_tab = f"{class_tab_base} - {date_str}"
    try:
        ws = sheet.worksheet(today_tab)
        # Clear the sheet if it already exists (reset for the day)
        ws.clear()
        ws.append_row([
            "name", "roll_no", "subject_name", "date", "day",
            "entry_time", "exit_time", "timestamp",
            "department_name", "semester_name", "group_name"
        ])
    except gspread.exceptions.WorksheetNotFound:
        ws = sheet.add_worksheet(title=today_tab, rows="1000", cols="20")
        ws.append_row([
            "name", "roll_no", "subject_name", "date", "day",
            "entry_time", "exit_time", "timestamp",
            "department_name", "semester_name", "group_name"
        ])

    # 3. Append today’s entry
    ws.append_row(row_data)



def get_or_create_class_sheet(dept_name: str, sem_name: str, grp_name: str):
    if sheet is None:
        return None
    tab_name = f"{dept_name} - {sem_name} - {grp_name}"
    try:
        ws = sheet.worksheet(tab_name)
    except gspread.exceptions.WorksheetNotFound:
        ws = sheet.add_worksheet(title=tab_name, rows="1000", cols="12")
        ws.append_row([
            "name", "roll_no", "subject_name", "date", "day",
            "entry_time", "timestamp",
            "department_name", "semester_name", "group_name"
        ])
    return ws

def export_row_to_google_sheets(row_values: list, dept_name: str, sem_name: str, grp_name: str):
    if internet_available and sheet is not None and GSHEETS_AVAILABLE:
        try:
            ws = get_or_create_class_sheet(dept_name, sem_name, grp_name)
            if ws:
                ws.append_row(row_values, value_input_option="USER_ENTERED")
        except Exception as e:
            print(f"⚠️ Google Sheets append failed: {e}")

def export_row_to_excel(date_str: str, row_values: list):
    try:
        wb, ws = ensure_backup_book_and_sheet(date_str)
        ws.append(row_values)
        wb.save(EXCEL_BACKUP_FILE)
    except Exception as e:
        print(f"⚠️ Excel export failed: {e}")

# ============================= MARK ATTENDANCE ================================
def mark_attendance(student_id: int):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    day_name = now.strftime("%A")
    time_str = now.strftime("%H:%M:%S")
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    time_hhmm = now.strftime("%H:%M")   # ✅ FIXED

    cursor.execute(
        """
        SELECT attendance_id FROM attendance
        WHERE student_id=? AND date=? AND exit_time IS NULL
        LIMIT 1
        """,
        (student_id, date_str),
    )
    existing = cursor.fetchone()

    if existing:
        attendance_id = existing[0]
        cursor.execute(
            "UPDATE attendance SET exit_time=?, timestamp=? WHERE attendance_id=?",
            (time_str, timestamp, attendance_id),
        )
        action = "exit"
    else:
        cursor.execute(
            """
            INSERT INTO attendance (student_id, date, day, entry_time, exit_time, timestamp)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (student_id, date_str, day_name, time_str, None, timestamp),
        )
        action = "entry"

    conn.commit()

    cursor.execute(
        """
        SELECT s.name, s.roll_no, s.semester_id, s.group_id, s.department_id
        FROM students s WHERE s.student_id=?
        """,
        (student_id,),
    )
    srow = cursor.fetchone()
    name, roll_no, semester_id, group_id, dept_id = srow if srow else ("", "", None, None, None)

    sem_name = cursor.execute("SELECT semester_name FROM semesters WHERE semester_id=?", (semester_id,)).fetchone()
    grp_name = cursor.execute("SELECT group_name FROM groups WHERE group_id=?", (group_id,)).fetchone()
    dept_name = cursor.execute("SELECT department_name FROM departments WHERE department_id=?", (dept_id,)).fetchone()

    sem_name = sem_name[0] if sem_name else ""
    grp_name = grp_name[0] if grp_name else ""
    dept_name = dept_name[0] if dept_name else ""

    row_values = [
        name, roll_no, "", 
        date_str, day_name,
        time_str if action == "entry" else "",
        time_str if action == "exit" else "",
        timestamp,
        dept_name, sem_name, grp_name
    ]

    export_row_to_excel(date_str, row_values)
    export_row_to_google_sheets(row_values, dept_name, sem_name, grp_name)

    print(f"📌 Attendance {action} for {name} ({roll_no}) at {time_str} — saved to {dept_name} - {sem_name} - {grp_name}")

# =============================
# UPDATE STUDENT
# =============================
def update_student():
    print("\n📝 Update student details")
    roll_no = input("Enter roll number of student to update: ").strip()

    cursor.execute("SELECT student_id, name FROM students WHERE roll_no=?", (roll_no,))
    row = cursor.fetchone()
    if not row:
        print("❌ Student not found.")
        return

    student_id, name = row
    print(f"Updating student: {name} ({roll_no})")

    new_sem = input("Enter new semester (leave blank to skip): ").strip()
    new_grp = input("Enter new group (leave blank to skip): ").strip()
    new_dept = input("Enter new department (leave blank to skip): ").strip()

    if new_sem:
        semester_id = get_or_create_semester(new_sem)
        cursor.execute("UPDATE students SET semester_id=? WHERE student_id=?", (semester_id, student_id))
    if new_grp:
        group_id = get_or_create_group(new_grp)
        cursor.execute("UPDATE students SET group_id=? WHERE student_id=?", (group_id, student_id))
    if new_dept:
        dept_id = get_or_create_department(new_dept)
        cursor.execute("UPDATE students SET department_id=? WHERE student_id=?", (dept_id, student_id))

    conn.commit()
    print("✅ Student updated successfully.")

# =============================
# ADD STUDENT

# =============================
def add_student():
    if not FACE_RECOGNITION_AVAILABLE:
        print("❌ Face recognition module is required for adding students with face encoding.")
        print("Install it with: pip install face-recognition")
        return
        
    print("\n🆕 Add a new student to the database")
    name = input("Enter student name: ").strip()
    roll_no = input("Enter roll number: ").strip()
    group_name = input("Enter group (e.g., G5/G8): ").strip()
    semester_name = input("Enter semester (e.g., sem4): ").strip()
    department_name = input("Enter department (e.g., CSE): ").strip()

    group_id = get_or_create_group(group_name)
    semester_id = get_or_create_semester(semester_name)
    department_id = get_or_create_department(department_name)

    cursor.execute(
        """
        INSERT OR IGNORE INTO students (name, roll_no, class_name, semester_id, group_id, department_id)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (name, roll_no, group_name, semester_id, group_id, department_id),
    )
    conn.commit()

    student_id = cursor.execute("SELECT student_id FROM students WHERE roll_no=?", (roll_no,)).fetchone()[0]

    student_folder = os.path.join(IMAGE_FOLDER, name)
    os.makedirs(student_folder, exist_ok=True)

    print("\nHow do you want to add student images?")
    print("1. Capture using webcam (recommended)")
    print("2. Provide your own images in folder (place files in {})".format(student_folder))
    choice = input("Enter choice (1/2): ").strip()

    if choice == "1":
        print("📸 Press SPACE to capture photos (3–4 angles). Press ESC to stop.")
        cap = open_camera(0)
        if cap is None:
            return  # early exit if camera not available
        photo_count = 0
        captured_any = False
        cv2.setWindowTitle("Attendance System", "Capture Student Face")
        while True:
            ret, frame = cap.read()
            if not ret:
                print("⚠️ Camera read failed.")
                break

            frame = maybe_rotate(frame)
            cv2.imshow("Attendance System", frame)

            key = cv2.waitKey(1) & 0xFF
            if key == 27:  # ESC
                break
            elif key == 32:  # SPACE
                photo_count += 1
                filename = f"{roll_no}_{photo_count}.png"
                path = os.path.join(student_folder, filename)
                cv2.imwrite(path, frame)
                print(f"✅ Saved {path}")
                captured_any = True
                if photo_count >= 4:
                    break
        cap.release()
        cv2.destroyAllWindows()
        if not captured_any:
            print("⚠️ No photos were captured. You may add files manually to:", student_folder)
    elif choice == "2":
        print("⚡ OK — place images into:", student_folder)
    else:
        print("⚠️ Invalid choice — skipping image capture.")

    # Load encodings from student folder
    added = 0
    for file in os.listdir(student_folder):
        path = os.path.join(student_folder, file)
        img = cv2.imread(path)
        if img is None:
            continue
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encs = face_recognition.face_encodings(rgb_img)
        if encs:
            known_face_encodings.append(encs[0])
            known_face_ids[len(known_face_encodings) - 1] = student_id
            added += 1

    print(f"✅ Student {name} ({roll_no}) added. {added} encodings loaded.")

# =============================
# LOAD STUDENTS (from images folder)
# =============================
def load_students():
    if not FACE_RECOGNITION_AVAILABLE:
        print("❌ Face recognition module is required for loading student face encodings.")
        return
        
    cursor.execute("SELECT student_id, name FROM students")
    students = cursor.fetchall()
    loaded = 0
    for student_id, name in students:
        student_folder = os.path.join(IMAGE_FOLDER, name)
        if not os.path.exists(student_folder):
            continue
        for file in os.listdir(student_folder):
            path = os.path.join(student_folder, file)
            img = cv2.imread(path)
            if img is None:
                continue
            rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encs = face_recognition.face_encodings(rgb_img)
            if encs:
                known_face_encodings.append(encs[0])
                known_face_ids[len(known_face_encodings) - 1] = student_id
                loaded += 1
    print(f"✅ Loaded {loaded} encodings for {len(students)} students.")

# =============================
# RECOGNITION LOOP
# =============================
def start_recognition():
    if not FACE_RECOGNITION_AVAILABLE:
        print("❌ Face recognition module is required for attendance recognition.")
        return
        
    if not known_face_encodings:
        print("⚠️ No known face encodings loaded. Run 'Load students' first.")
        return

    cap = open_camera(0)
    if cap is None:
        return

    cv2.setWindowTitle("Attendance System", "Attendance System")

    while True:
        success, frame = cap.read()
        if not success:
            print("⚠️ Camera frame read failed.")
            break

        frame = maybe_rotate(frame)

        # Process smaller image for speed
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small)
        face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)

            name = "Unknown"
            student_id = None

            if True in matches:
                best_idx = int(np.argmin(face_distances))
                student_id = known_face_ids.get(best_idx)

                if student_id:
                    now = datetime.now()
                    last = _last_seen.get(student_id)
                    if (last is None) or ((now - last) > timedelta(seconds=COOLDOWN_SECONDS)):
                        mark_attendance(student_id)
                        _last_seen[student_id] = now

                    cursor.execute("SELECT name FROM students WHERE student_id=?", (student_id,))
                    rr = cursor.fetchone()
                    name = rr[0] if rr else "Unknown"
            else:
                # save unknown face crop at original scale (map back to full frame)
                top, right, bottom, left = [v * 4 for v in face_location]
                today_folder = os.path.join(UNKNOWN_FOLDER, datetime.now().strftime("%Y-%m-%d"))
                os.makedirs(today_folder, exist_ok=True)
                filename = f"unknown_{datetime.now().strftime('%H%M%S')}.png"
                path = os.path.join(today_folder, filename)
                try:
                    cv2.imwrite(path, frame[top:bottom, left:right])
                except Exception:
                    pass

            # draw rect + label on full frame (map back to full frame)
            top, right, bottom, left = [v * 4 for v in face_location]
            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

        cv2.imshow("Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == 27:  # ESC
            break

    cap.release()
    cv2.destroyAllWindows()


# =============================
# UPDATE STUDENT
# =============================
# The `update_student` function is defined earlier (keeps interactive flow).

# Stubs for additional menu options so the script runs on Windows/macOS without
# NameError if those features aren't implemented yet.
def manage_subjects():
    print("\n📚 Manage Subjects — not implemented in this build.\n")


def manage_timetable():
    print("\n🗓️ Manage Timetable — not implemented in this build.\n")

# # =============================
# # MAIN MENU
# # =============================

# if __name__ == "__main__":
#     try:
#         while True:
#             print("\n==== FACE RECOGNITION ATTENDANCE ====")
#             print("1. Add new student")
#             print("2. Load students & start recognition")
#             print("3. Update student")
#             print("4. Exit")
#             choice = input("Enter choice: ").strip()

#             if choice == "1":
#                 add_student()
#             elif choice == "2":
#                 load_students()
#                 start_recognition()
#             elif choice == "3":
#                 update_student()
#             elif choice == "4":
#                 print("Exiting...")
#                 break
#             else:
#                 print("❌ Invalid choice")
#     finally:
#         conn.close()


 # MAIN MENU
# =============================
def show_system_status():
    print("\n=== SYSTEM STATUS ===")
    print(f"🖥️  Platform: {platform.system()}")
    print(f"📹 OpenCV: ✅ v{cv2.__version__}")
    print(f"🔍 Face Recognition: {'✅ Available' if FACE_RECOGNITION_AVAILABLE else '❌ Not Available'}")
    print(f"📊 Google Sheets: {'✅ Available' if GSHEETS_AVAILABLE else '❌ Not Available'}")
    print(f"🌐 Internet: {'✅ Connected' if internet_available else '❌ Offline'}")
    
    if not FACE_RECOGNITION_AVAILABLE:
        print("\n⚠️  To enable face recognition features:")
        print("   pip install face-recognition")
        print("   (Requires Visual Studio Build Tools on Windows)")

def test_camera_basic():
    print("\n📸 Testing Camera...")
    cap = open_camera(0)
    if cap is None:
        print("❌ Camera could not be opened!")
        return
    
    print("Camera opened successfully. Press ESC to close.")
    print("📹 A camera window should appear...")
    
    try:
        while True:
            success, frame = cap.read()
            if not success:
                print("❌ Failed to read from camera")
                break
            
            frame = maybe_rotate(frame)
            cv2.putText(frame, "Camera Test - Press ESC to exit", (10, 30), 
                       cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            
            cv2.imshow("Camera Test", frame)
            
            key = cv2.waitKey(1) & 0xFF
            if key == 27:  # ESC
                print("🎯 ESC pressed - closing camera")
                break
                
    except Exception as e:
        print(f"❌ Camera test error: {e}")
    finally:
        cap.release()
        cv2.destroyAllWindows()
        print("✅ Camera test completed")

def export_to_excel():
    """Export attendance data to Excel file"""
    try:
        # Get attendance data with student names
        cursor.execute("""
            SELECT s.name, s.roll_no, s.class_name, a.date, a.day, a.entry_time, a.timestamp
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            ORDER BY a.timestamp DESC
        """)
        
        attendance_data = cursor.fetchall()
        
        if not attendance_data:
            print("⚠️  No attendance data to export")
            return
        
        # Create DataFrame
        df = pd.DataFrame(attendance_data, columns=['Name', 'Roll_No', 'Class', 'Date', 'Day', 'Entry_Time', 'Timestamp'])
        
        # Generate filename with current date
        today = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"attendance_{today}.xlsx"
        
        # Export to Excel
        df.to_excel(filename, index=False)
        print(f"📊 Attendance exported to: {filename}")
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")

def capture_student_images(student_name, roll_no):
    """Capture multiple images of a student for enrollment"""
    print(f"\n📸 Capturing images for {student_name} (Roll: {roll_no})")
    
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("❌ Camera could not be opened!")
        return False
    
    # Create directory for student images
    student_dir = os.path.join("images", f"{student_name}")
    os.makedirs(student_dir, exist_ok=True)
    
    print("📹 Camera opened. Instructions:")
    print("   - Position your face clearly in the camera")
    print("   - Press SPACE to capture an image")
    print("   - Capture 3-4 different angles/expressions")
    print("   - Press ESC when done")
    
    image_count = 0
    target_images = 4
    
    try:
        while image_count < target_images:
            success, frame = cap.read()
            if not success:
                print("❌ Failed to read from camera")
                break
            
            frame = maybe_rotate(frame)
            
            # Add text overlay
            cv2.putText(frame, f"Capturing: {student_name}", (10, 30), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, f"Images: {image_count}/{target_images}", (10, 60), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, "SPACE = Capture, ESC = Done", (10, 90), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 2)
            
            cv2.imshow("Student Image Capture", frame)
            
            key = cv2.waitKey(1) & 0xFF
            if key == 32:  # SPACE
                # Save the image
                image_filename = f"{roll_no}_{image_count + 1}.png"
                image_path = os.path.join(student_dir, image_filename)
                cv2.imwrite(image_path, frame)
                print(f"✅ Image {image_count + 1} captured: {image_filename}")
                image_count += 1
                
            elif key == 27:  # ESC
                break
                
    except Exception as e:
        print(f"❌ Error during image capture: {e}")
        return False
    finally:
        cap.release()
        cv2.destroyAllWindows()
    
    if image_count > 0:
        print(f"✅ Successfully captured {image_count} images for {student_name}")
        return True
    else:
        print("⚠️  No images were captured")
        return False

def load_student_images():
    """Load all student images and create a recognition database"""
    print("\n📂 Loading student images...")
    
    student_images = {}
    images_dir = "images"
    
    if not os.path.exists(images_dir):
        print(f"❌ Images directory '{images_dir}' not found")
        return {}
    
    loaded_count = 0
    for student_folder in os.listdir(images_dir):
        student_path = os.path.join(images_dir, student_folder)
        if os.path.isdir(student_path):
            student_images[student_folder] = []
            
            for image_file in os.listdir(student_path):
                if image_file.lower().endswith(('.png', '.jpg', '.jpeg')):
                    image_path = os.path.join(student_path, image_file)
                    try:
                        image = cv2.imread(image_path)
                        if image is not None:
                            student_images[student_folder].append(image)
                            loaded_count += 1
                    except Exception as e:
                        print(f"⚠️  Could not load {image_path}: {e}")
    
    print(f"✅ Loaded {loaded_count} images for {len(student_images)} students")
    return student_images

def compare_faces_basic(captured_frame, student_images):
    """Basic face comparison using template matching (without face_recognition library)"""
    matches = []
    
    # Convert captured frame to grayscale
    gray_frame = cv2.cvtColor(captured_frame, cv2.COLOR_BGR2GRAY)
    
    # Use Haar cascade for face detection
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces_in_frame = face_cascade.detectMultiScale(gray_frame, 1.3, 5)
    
    if len(faces_in_frame) == 0:
        return matches
    
    for student_name, images in student_images.items():
        max_similarity = 0
        
        for student_image in images:
            try:
                # Convert student image to grayscale
                gray_student = cv2.cvtColor(student_image, cv2.COLOR_BGR2GRAY)
                
                # Detect face in student image
                student_faces = face_cascade.detectMultiScale(gray_student, 1.3, 5)
                
                if len(student_faces) > 0:
                    # Get the first face from both images
                    (x1, y1, w1, h1) = faces_in_frame[0]
                    (x2, y2, w2, h2) = student_faces[0]
                    
                    # Extract face regions
                    face_frame = gray_frame[y1:y1+h1, x1:x1+w1]
                    face_student = gray_student[y2:y2+h2, x2:x2+w2]
                    
                    # Resize to same size for comparison
                    face_student = cv2.resize(face_student, (w1, h1))
                    
                    # Calculate similarity using template matching
                    result = cv2.matchTemplate(face_frame, face_student, cv2.TM_CCOEFF_NORMED)
                    similarity = cv2.minMaxLoc(result)[1]
                    
                    if similarity > max_similarity:
                        max_similarity = similarity
                        
            except Exception as e:
                continue
        
        if max_similarity > 0.6:  # Threshold for face match
            matches.append((student_name, max_similarity))
    
    # Sort by similarity score
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches

def add_student_basic():
    """Add student with image capture option"""
    print("\n🆕 Add a new student to the database")
    
    name = input("Enter student name: ").strip()
    if not name:
        print("❌ Name cannot be empty")
        return
        
    roll_no = input("Enter roll number: ").strip()
    if not roll_no:
        print("❌ Roll number cannot be empty")
        return
        
    group_name = input("Enter group (e.g., G5/G8): ").strip()
    semester_name = input("Enter semester (e.g., sem4): ").strip()
    department_name = input("Enter department (e.g., CSE): ").strip()

    # Ask if user wants to capture images
    capture_images = input("\n📸 Do you want to capture student images? (y/n): ").strip().lower()
    
    try:
        # Get or create related records
        group_id = get_or_create_group(group_name) if group_name else 1
        semester_id = get_or_create_semester(semester_name) if semester_name else 1
        department_id = get_or_create_department(department_name) if department_name else 1

        # Insert student
        cursor.execute(
            """
            INSERT OR IGNORE INTO students (name, roll_no, class_name, semester_id, group_id, department_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (name, roll_no, group_name, semester_id, group_id, department_id)
        )
        
        if cursor.rowcount > 0:
            student_id = cursor.lastrowid
            conn.commit()
            print(f"✅ Student {name} (Roll: {roll_no}) added successfully!")
            print(f"📁 Student ID: {student_id}")
            
            # Capture images if requested
            if capture_images == 'y':
                if capture_student_images(name, roll_no):
                    print("📸 Student images captured and saved!")
                else:
                    print("⚠️  Images not captured, but student data is saved")
            else:
                print("📝 Student added without images")
                
        else:
            print(f"⚠️  Student {name} with roll number {roll_no} already exists")
            
    except Exception as e:
        print(f"❌ Error adding student: {e}")

def start_manual_attendance():
    """Manual attendance marking without face recognition"""
    print("\n📝 Manual Attendance System")
    print("Choose attendance method:")
    print("1. Manual entry (select students)")
    print("2. Camera-based verification (using student images)")
    
    method = input("Enter choice (1 or 2): ").strip()
    
    if method == "2":
        start_image_attendance()
        return
    
    print("👤 Using manual entry method")
    
    # Show available students
    cursor.execute("SELECT student_id, name, roll_no, class_name FROM students ORDER BY name")
    students = cursor.fetchall()
    
    if not students:
        print("❌ No students found in database")
        print("💡 Add students first using option 3")
        return
    
    print("\n📋 Available Students:")
    for i, (student_id, name, roll_no, class_name) in enumerate(students, 1):
        print(f"{i}. {name} (Roll: {roll_no}) - {class_name}")
    
    print("\n🎯 Mark Attendance:")
    print("Enter student numbers separated by commas (e.g., 1,3,5)")
    print("Or type 'all' to mark all present, 'q' to quit")
    
    choice = input("\nYour choice: ").strip().lower()
    
    if choice == 'q':
        return
    elif choice == 'all':
        selected_indices = list(range(len(students)))
    else:
        try:
            # Parse comma-separated numbers
            selected_indices = [int(x.strip()) - 1 for x in choice.split(',') if x.strip().isdigit()]
            selected_indices = [i for i in selected_indices if 0 <= i < len(students)]
        except:
            print("❌ Invalid input format")
            return
    
    if not selected_indices:
        print("❌ No valid students selected")
        return
    
    # Mark attendance
    current_time = datetime.now()
    marked_count = 0
    
    for i in selected_indices:
        student_id, name, roll_no, class_name = students[i]
        
        try:
            # Check if student already marked today
            today_date = current_time.strftime('%Y-%m-%d')
            cursor.execute(
                "SELECT COUNT(*) FROM attendance WHERE student_id = ? AND date = ?",
                (student_id, today_date)
            )
            
            if cursor.fetchone()[0] == 0:
                # Student not marked today, add attendance
                cursor.execute(
                    """
                    INSERT INTO attendance (student_id, date, day, entry_time, timestamp)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (student_id, today_date, current_time.strftime('%A'), 
                     current_time.strftime('%H:%M:%S'), current_time.strftime('%Y-%m-%d %H:%M:%S'))
                )
                print(f"✅ {name} (Roll: {roll_no}) marked present")
                marked_count += 1
            else:
                print(f"⚠️  {name} (Roll: {roll_no}) already marked today")
                
        except Exception as e:
            print(f"❌ Error marking {name}: {e}")
    
    conn.commit()
    print(f"\n🎉 Attendance marked for {marked_count} students")
    
    # Export to Excel
    try:
        export_to_excel()
        print("📊 Attendance exported to Excel")
    except Exception as e:
        print(f"⚠️  Excel export failed: {e}")

def start_image_attendance():
    """Camera-based attendance using student images"""
    print("\n📸 Camera-Based Attendance System")
    
    # Load student images
    student_images = load_student_images()
    if not student_images:
        print("❌ No student images found!")
        print("💡 Add student images first using option 3")
        return
    
    print(f"✅ Loaded images for {len(student_images)} students")
    
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("❌ Camera could not be opened!")
        return
    
    print("\n📹 Camera opened. Instructions:")
    print("   - Show your face to the camera")
    print("   - Press SPACE to capture and verify")
    print("   - Press ESC to finish attendance")
    
    marked_students = set()
    
    try:
        while True:
            success, frame = cap.read()
            if not success:
                print("❌ Failed to read from camera")
                break
            
            frame = maybe_rotate(frame)
            
            # Add text overlay
            cv2.putText(frame, "Attendance System - SPACE to verify", (10, 30), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, f"Marked: {len(marked_students)} students", (10, 60), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, "ESC = Finish", (10, 90), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 2)
            
            cv2.imshow("Attendance - Face Verification", frame)
            
            key = cv2.waitKey(1) & 0xFF
            if key == 32:  # SPACE
                # Verify face
                matches = compare_faces_basic(frame, student_images)
                
                if matches:
                    best_match = matches[0]
                    student_name = best_match[0]
                    confidence = best_match[1]
                    
                    if student_name not in marked_students:
                        # Mark attendance
                        try:
                            # Get student ID from database
                            cursor.execute("SELECT student_id FROM students WHERE name = ?", (student_name,))
                            result = cursor.fetchone()
                            
                            if result:
                                student_id = result[0]
                                current_time = datetime.now()
                                today_date = current_time.strftime('%Y-%m-%d')
                                
                                # Check if already marked today
                                cursor.execute(
                                    "SELECT COUNT(*) FROM attendance WHERE student_id = ? AND date = ?",
                                    (student_id, today_date)
                                )
                                
                                if cursor.fetchone()[0] == 0:
                                    # Not marked today, add attendance
                                    cursor.execute(
                                        """
                                        INSERT INTO attendance (student_id, date, day, entry_time, timestamp)
                                        VALUES (?, ?, ?, ?, ?)
                                        """,
                                        (student_id, today_date, current_time.strftime('%A'),
                                         current_time.strftime('%H:%M:%S'), current_time.strftime('%Y-%m-%d %H:%M:%S'))
                                    )
                                    marked_students.add(student_name)
                                    conn.commit()
                                    print(f"✅ {student_name} marked present (Confidence: {confidence:.2f})")
                                else:
                                    print(f"⚠️  {student_name} already marked today")
                            else:
                                print(f"❌ Student {student_name} not found in database")
                                
                        except Exception as e:
                            print(f"❌ Error marking {student_name}: {e}")
                    else:
                        print(f"⚠️  {student_name} already marked present today")
                else:
                    print("❌ No matching student found")
                    
            elif key == 27:  # ESC
                break
                
    except Exception as e:
        print(f"❌ Error during attendance: {e}")
    finally:
        cap.release()
        cv2.destroyAllWindows()
    
    print(f"\n🎉 Attendance session completed!")
    print(f"📊 Total students marked present: {len(marked_students)}")
    
    if marked_students:
        print("✅ Students marked present:")
        for student in marked_students:
            print(f"   - {student}")
        
        # Export to Excel
        try:
            export_to_excel()
            print("📊 Attendance exported to Excel")
        except Exception as e:
            print(f"⚠️  Excel export failed: {e}")

def update_student_basic():
    """Update student information without face recognition"""
    print("\n✏️ Update Student Information (Basic Mode)")
    
    # Show available students
    cursor.execute("SELECT student_id, name, roll_no, class_name FROM students ORDER BY name")
    students = cursor.fetchall()
    
    if not students:
        print("❌ No students found in database")
        return
    
    print("\n📋 Available Students:")
    for i, (student_id, name, roll_no, class_name) in enumerate(students, 1):
        print(f"{i}. {name} (Roll: {roll_no}) - {class_name}")
    
    try:
        choice = int(input("\nSelect student number to update (0 to cancel): "))
        if choice == 0:
            return
        if choice < 1 or choice > len(students):
            print("❌ Invalid selection")
            return
        
        student_id, current_name, current_roll, current_class = students[choice - 1]
        
        print(f"\n📝 Updating: {current_name} (Roll: {current_roll})")
        print("💡 Press Enter to keep current value")
        
        # Get new values
        new_name = input(f"Name [{current_name}]: ").strip()
        new_roll = input(f"Roll Number [{current_roll}]: ").strip()
        new_class = input(f"Class [{current_class}]: ").strip()
        
        # Use current values if empty
        new_name = new_name if new_name else current_name
        new_roll = new_roll if new_roll else current_roll
        new_class = new_class if new_class else current_class
        
        # Update database
        cursor.execute(
            """
            UPDATE students 
            SET name = ?, roll_no = ?, class_name = ?
            WHERE student_id = ?
            """,
            (new_name, new_roll, new_class, student_id)
        )
        
        conn.commit()
        print(f"✅ Student updated successfully!")
        print(f"📁 New details: {new_name} (Roll: {new_roll}) - {new_class}")
        
    except ValueError:
        print("❌ Invalid input")
    except Exception as e:
        print(f"❌ Error updating student: {e}")

if __name__ == "__main__":
    print("🚀 Starting Face Recognition Attendance System...")
    show_system_status()
    
    try:
        while True:
            print("\n==== FACE RECOGNITION ATTENDANCE ====")
            print("1. Test Camera")
            print("2. Show System Status")
            
            if FACE_RECOGNITION_AVAILABLE:
                print("3. Add new student (with face recognition)")
                print("4. Start face recognition attendance")
                print("5. Update student details")
            else:
                print("3. Add new student (with image capture)")
                print("4. Attendance system (manual/camera-based)")
                print("5. Update student details")
            
            print("6. Manage Subjects")
            print("7. Manage Timetable")
            print("8. Exit")
            choice = input("\nEnter choice: ").strip()

            if choice == "1":
                test_camera_basic()
            elif choice == "2":
                show_system_status()
            elif choice == "3":
                if FACE_RECOGNITION_AVAILABLE:
                    add_student()
                else:
                    add_student_basic()
            elif choice == "4":
                if FACE_RECOGNITION_AVAILABLE:
                    load_students()
                    start_recognition()
                else:
                    start_manual_attendance()
            elif choice == "5":
                if FACE_RECOGNITION_AVAILABLE:
                    update_student()
                else:
                    update_student_basic()
            elif choice == "6":
                manage_subjects()
            elif choice == "7":
                manage_timetable()
            elif choice == "8":
                print("Exiting...")
                break
            else:
                print("❌ Invalid choice. Please select 1-8.")
    finally:
        conn.close()

